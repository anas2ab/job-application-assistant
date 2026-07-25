from datetime import date
from pathlib import Path
from typing import Set

from openpyxl import load_workbook

from config import TRACKER_PATH
from models import Job


APPLICATION_SHEET = "Applications"


def normalize_url(url: str) -> str:
    return url.strip().lower().rstrip("/")


def get_existing_urls() -> Set[str]:
    path = Path(TRACKER_PATH)

    if not path.exists():
        raise FileNotFoundError(
            f"Tracker not found: {path.resolve()}"
        )

    workbook = load_workbook(path)
    worksheet = workbook[APPLICATION_SHEET]

    existing_urls: Set[str] = set()

    # Job URL is column F
    for row_number in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=6).value

        if value:
            existing_urls.add(normalize_url(str(value)))

    workbook.close()
    return existing_urls


def find_first_available_row(worksheet) -> int:
    # Company is column B.
    for row_number in range(2, worksheet.max_row + 1):
        company = worksheet.cell(
            row=row_number,
            column=2,
        ).value

        if company is None or str(company).strip() == "":
            return row_number

    return worksheet.max_row + 1


def add_jobs(jobs: list[Job]) -> int:
    path = Path(TRACKER_PATH)

    if not path.exists():
        raise FileNotFoundError(
            f"Tracker not found: {path.resolve()}"
        )

    workbook = load_workbook(path)
    worksheet = workbook[APPLICATION_SHEET]

    existing_urls = get_existing_urls()
    added = 0

    for job in jobs:
        normalized_url = normalize_url(job.url)

        if not normalized_url:
            continue

        if normalized_url in existing_urls:
            continue

        row = find_first_available_row(worksheet)

        # Preserve formula cells already present in the template.
        worksheet.cell(row=row, column=2, value=job.company)
        worksheet.cell(row=row, column=3, value=job.title)
        worksheet.cell(row=row, column=4, value=job.location)
        worksheet.cell(row=row, column=5, value=job.work_type)
        worksheet.cell(row=row, column=6, value=job.url)
        worksheet.cell(row=row, column=7, value=job.source)
        worksheet.cell(row=row, column=8, value=date.today())
        worksheet.cell(row=row, column=10, value="Discovered")
        worksheet.cell(
            row=row,
            column=11,
            value="High" if job.score >= 80 else "Medium",
        )
        worksheet.cell(row=row, column=12, value=job.score)
        worksheet.cell(
            row=row,
            column=19,
            value="Review job description",
        )
        worksheet.cell(
            row=row,
            column=21,
            value=job.description[:3000],
        )

        existing_urls.add(normalized_url)
        added += 1

    workbook.save(path)
    workbook.close()

    return added